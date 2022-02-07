#se utiliza la librería openpyxl para trabajar con archivos xlsx
from openpyxl import load_workbook
import os
#lectura del archvio de excel
libro = load_workbook('lista-so.xlsx')

hoja = libro.active

hojax = libro['asistencia'] #se establece en la hoja de asistencia del archivo

#lee el directorio cedulas y obtiene cada archivo de texto (imagen escaneada) de cada fecha en una lista
for ruta, dir, files in os.walk('cedulas', topdown = False):
	archivos = files

cedulas_fecha = sorted(archivos) #ordena la lista por fecha

#abre el archivo de fechas como lectura y cada línea se agrega a una lista 
with open('fechas.txt','r') as archivo_f:
	fecha = [linea.strip("\n") for linea in archivo_f]

fecha = list(set(fecha)) #con set elimina las fechas (año, mes, dia) repetidas. (no permite elementos duplicados)

#recorre la lista fecha por su longitud y añade a cada columna cada elemento de la lista
for i in range(0,len(fecha)):

	hojax.cell(row=1,column=(i+4), value = fecha[i])
	
	#abre cada archivo de texto del directorio de las cédulas. se separa por elementos pares e impares
	#debido a que existe dos fotos para una misma fecha.
	with open("cedulas/" + cedulas_fecha[i*2],'r') as archivo_c1:
  	      ced1 = [linea.strip("\n") for linea in archivo_c1]

	with open("cedulas/" + cedulas_fecha[i*2+1],'r') as archivo_c2:
              ced2 = [linea.strip("\n") for linea in archivo_c2]

	#recorre las filas del documento xlsx donde están todas las cédulas
	for filas in hojax["B2":"B46"]:
		for celdas in filas:

			cedula_estudiante = celdas.value #obtiene el valor de cada celda (fila de cedulas)
			n_fila = celdas.row #obtiene el numero de fila de la celda seleccionada

			#evalúa si cada cédula en el archivo excel tiene aparición en la lista de cada foto escaneada
			if cedula_estudiante in ced1:
				hojax.cell(row=n_fila, column=(i+4), value = "S")

			if cedula_estudiante in ced2:
				hojax.cell(row=n_fila, column=(i+4), value = "S")

			else:
				hojax.cell(row=n_fila, column=(i+4), value = "N")



libro.save("lista-so.xlsx") #guarda los cambios efectuados del codigo al archivo .xlsx
